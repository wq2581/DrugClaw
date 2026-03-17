"""
27_ATC_DDD  –  WHO ATC/DDD Classification System Query Skill
Category: Drug-centric | Type: DB + API | Subcategory: Drug Ontology/Terminology

Data source (local):
    ATC_DDD_new and alterations 2026_final.xlsx
    Columns: ATC code, ATC level name, New DDD, Unit, Adm.route, Note

Data source (remote – RxNav/RxClass):
    https://rxnav.nlm.nih.gov/REST/rxclass/
"""

import re
import csv
import json
import urllib.request
import urllib.parse
import urllib.error
from typing import Optional

# ── Paths ────────────────────────────────────────────────────────────
DATA_DIR  = "/blue/qsong1/wang.qing/AgentLLM/Survey100/resources_metadata/drug_ontology/ATC_DDD"
XLSX_PATH = f"{DATA_DIR}/ATC_DDD_new and alterations 2026_final.xlsx"

RXNAV_BASE = "https://rxnav.nlm.nih.gov/REST"

# ── Pattern detection ────────────────────────────────────────────────
_RE_ATC = re.compile(r"^[A-Z]\d{2}(?:[A-Z]{1,2}\d{0,2})?$", re.IGNORECASE)

_cache: dict = {}


# ── Helpers ──────────────────────────────────────────────────────────
def _get_json(url: str, timeout: int = 20) -> Optional[dict]:
    """GET JSON from URL; return None on HTTP error."""
    try:
        req = urllib.request.Request(url, headers={"Accept": "application/json"})
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            return json.loads(resp.read())
    except (urllib.error.HTTPError, urllib.error.URLError, Exception) as exc:
        print(f"  [WARN] {exc}")
        return None


def _detect_type(entity: str) -> str:
    """Return 'atc_code' or 'drug_name'."""
    return "atc_code" if _RE_ATC.match(entity.strip()) else "drug_name"


# ── Local Excel loader ───────────────────────────────────────────────
def load_local(path: str = XLSX_PATH) -> list[dict]:
    """Load ATC/DDD Excel into a list of dicts.  Cached after first call."""
    if "local" in _cache:
        return _cache["local"]
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception:
        # Fallback: try pandas
        import pandas as pd
        df = pd.read_excel(path)
        records = df.to_dict(orient="records")
        _cache["local"] = records
        return records

    # Normalise header
    header_map = {
        "atc code": "atc_code",
        "atc level name": "atc_level_name",
        "new ddd": "new_ddd",
        "unit": "unit",
        "adm.route": "adm_route",
        "note": "note",
    }
    raw_header = [str(c).strip().lower() if c else "" for c in rows[0]]
    header = [header_map.get(h, h) for h in raw_header]
    records = []
    for row in rows[1:]:
        rec = {}
        for i, val in enumerate(row):
            if i < len(header):
                rec[header[i]] = val if val is not None else ""
        if rec.get("atc_code"):
            records.append(rec)
    _cache["local"] = records
    return records


def _search_local(entity: str, records: list[dict]) -> list[dict]:
    """Search local records by ATC code prefix or name substring."""
    ent = entity.strip()
    etype = _detect_type(ent)
    if etype == "atc_code":
        code_up = ent.upper()
        return [r for r in records if str(r.get("atc_code", "")).upper().startswith(code_up)]
    else:
        low = ent.lower()
        return [r for r in records if low in str(r.get("atc_level_name", "")).lower()]


# ── RxNav API queries ────────────────────────────────────────────────
def _rxnav_drug_to_atc(drug_name: str) -> list[dict]:
    """Drug name → list of ATC class dicts via RxClass."""
    url = (f"{RXNAV_BASE}/rxclass/class/byDrugName.json"
           f"?drugName={urllib.parse.quote(drug_name)}&relaSource=ATC")
    data = _get_json(url)
    if not data:
        return []
    infos = data.get("rxclassDrugInfoList", {}).get("rxclassDrugInfo", [])
    seen, results = set(), []
    for info in infos:
        cls = info.get("rxclassMinConceptItem", {})
        code = cls.get("classId", "")
        if code and code not in seen:
            seen.add(code)
            results.append({
                "atc_code": code,
                "atc_level_name": cls.get("className", ""),
                "source": "RxNav",
            })
    return results


def _rxnav_atc_to_drugs(atc_code: str) -> list[dict]:
    """ATC code → list of member drug dicts via RxClass."""
    params = urllib.parse.urlencode({"classId": atc_code, "relaSource": "ATC"})
    url = f"{RXNAV_BASE}/rxclass/classMembers.json?{params}"
    data = _get_json(url)
    if not data:
        return []
    members = data.get("drugMemberGroup", {}).get("drugMember", [])
    seen, results = set(), []
    for m in members:
        c = m.get("minConcept", {})
        name = c.get("name", "")
        if name and name not in seen:
            seen.add(name)
            results.append({
                "drug_name": name,
                "rxcui": c.get("rxcui", ""),
                "source": "RxNav",
            })
    return results


# ── Public interface ─────────────────────────────────────────────────
def search(entity: str, *, use_api: bool = True) -> dict:
    """Query ATC/DDD for a single entity (ATC code or drug name).

    Returns dict with keys: entity, entity_type, local_hits, api_hits.
    """
    ent = entity.strip()
    etype = _detect_type(ent)

    # Local lookup
    try:
        records = load_local()
    except Exception:
        records = []
    local_hits = _search_local(ent, records) if records else []

    # API lookup
    api_hits: list[dict] = []
    if use_api:
        try:
            if etype == "atc_code":
                api_hits = _rxnav_atc_to_drugs(ent.upper())
            else:
                api_hits = _rxnav_drug_to_atc(ent)
        except Exception as exc:
            print(f"  [WARN] API error: {exc}")

    return {
        "entity": ent,
        "entity_type": etype,
        "local_hits": local_hits,
        "api_hits": api_hits,
    }


def search_batch(entities: list[str], *, use_api: bool = True) -> dict[str, dict]:
    """Query multiple entities. Returns {entity: search_result}."""
    return {e: search(e, use_api=use_api) for e in entities}


def summarize(result: dict) -> str:
    """Return a concise LLM-readable text summary of a search result."""
    ent   = result["entity"]
    etype = result["entity_type"]
    local = result.get("local_hits", [])
    api   = result.get("api_hits", [])
    lines = [f"### ATC/DDD result for '{ent}' (detected: {etype})"]

    if local:
        lines.append(f"**Local DDD records ({len(local)}):**")
        for r in local:
            ddd_str = f"{r.get('new_ddd', '')} {r.get('unit', '')} {r.get('adm_route', '')}".strip()
            note = f" ({r['note']})" if r.get("note") else ""
            lines.append(f"  {r.get('atc_code','')}  {r.get('atc_level_name','')}"
                         f"  DDD={ddd_str}{note}")

    if api:
        if etype == "atc_code":
            lines.append(f"**Member drugs via RxNav ({len(api)}):**")
            for r in api[:15]:
                lines.append(f"  {r['drug_name']} (RxCUI:{r.get('rxcui','')})")
            if len(api) > 15:
                lines.append(f"  ... and {len(api)-15} more")
        else:
            lines.append(f"**ATC classifications via RxNav ({len(api)}):**")
            for r in api:
                lines.append(f"  {r['atc_code']}  {r['atc_level_name']}")

    if not local and not api:
        lines.append("No results found.")
    return "\n".join(lines)


def to_json(result: dict) -> str:
    """Return JSON string of search result."""
    return json.dumps(result, indent=2, ensure_ascii=False, default=str)


# ── CLI demo ─────────────────────────────────────────────────────────
if __name__ == "__main__":
    # --- Single entity: drug name ---
    r = search("aspirin")
    print(summarize(r))

    # --- Single entity: ATC code ---
    r = search("N02BA")
    print(summarize(r))

    # --- Batch search ---
    batch = search_batch(["metformin", "A10BA02", "C09XX01"])
    for ent, res in batch.items():
        print(summarize(res))
        print()

    # --- JSON output ---
    r = search("sparsentan")
    print(to_json(r))