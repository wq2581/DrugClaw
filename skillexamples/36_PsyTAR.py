"""
PsyTAR — Psychiatric Treatment Adverse Reaction Corpus
Category: Drug-centric | Type: Dataset (local XLSX) | Subcategory: Drug NLP / ADR
Paper: https://doi.org/10.1016/j.dib.2019.103838
License: CC BY 4.0

891 patient reviews for Zoloft, Lexapro, Cymbalta, Effexor XR from
askapatient.com.  6 009 sentences annotated for ADR / WD / SSI / DI / EF /
INF, with entities extracted and mapped to UMLS + SNOMED CT.

Access method: Local Excel file (openpyxl).

Sheets
------
Sample              Original posts (rating, indication, side-effect, comment …)
Sentence_Labeling   6 009 sentences with binary labels (ADR/WD/EF/INF/SSI/DI/Findings/Others)
ADR_Identified      Extracted ADR entities per sentence (ADR1, ADR2 …)
WD_Identified       Extracted WD entities
SSI_Identified      Extracted SSI entities
DI_Identified       Extracted DI entities
ADR_Mapped          ADR → UMLS1/UMLS2/SNOMED-CT + severity qualifiers
WD_Mapped           WD  → UMLS/SNOMED + qualifiers
SSI_Mapped          SSI → UMLS/SNOMED + qualifiers
DI_Mapped           DI  → UMLS/SNOMED + qualifiers
"""

import os, re, json

# ── path ────────────────────────────────────────────────────────────
DATA_PATH = os.environ.get(
    "PSYTAR_XLSX",
    "/blue/qsong1/wang.qing/AgentLLM/Survey100/"
    "resources_metadata/drug_nlp/PsyTAR/PsyTAR_dataset.xlsx",
)

# ── lazy cache ──────────────────────────────────────────────────────
_cache: dict[str, list[dict]] = {}   # sheet_name -> [row_dict …]

# ── constants ───────────────────────────────────────────────────────
_DRUG_ALIAS = {
    "sertraline": "zoloft", "escitalopram": "lexapro",
    "duloxetine": "cymbalta", "venlafaxine": "effexor",
    "effexor xr": "effexor",
}
_KNOWN_DRUGS = {"zoloft", "lexapro", "cymbalta", "effexor"}


# ====================================================================
#  Loading
# ====================================================================

def _load(path: str = DATA_PATH):
    if _cache:
        return
    try:
        import openpyxl
    except ImportError:
        raise ImportError("pip install openpyxl")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        hdrs = [str(h).strip() if h is not None else f"col_{i}"
                for i, h in enumerate(rows[0])]
        _cache[name] = [
            {h: (v if v is not None else "") for h, v in zip(hdrs, r)}
            for r in rows[1:]
        ]
    wb.close()


# ====================================================================
#  Helpers
# ====================================================================

def _n(x) -> str:
    """Normalise to lower-case stripped string."""
    return str(x).strip().lower()


def _drug(name: str) -> str:
    """Canonicalise drug name (generic → brand)."""
    n = _n(name)
    return _DRUG_ALIAS.get(n, n)


def _resolve(sheet: str) -> str | None:
    """Fuzzy-match user sheet name to actual sheet."""
    sl = sheet.lower()
    for s in _cache:
        if sl == s.lower() or sl in s.lower():
            return s
    return None


def _entity_type(entity: str) -> str:
    """Route: UMLS CUI | drug name | free text."""
    e = entity.strip()
    if re.match(r"^C\d{4,}$", e, re.I):
        return "cui"
    if _drug(e) in _KNOWN_DRUGS:
        return "drug"
    return "text"


# ── Column-aware matchers ───────────────────────────────────────────

def _match_drug(row: dict, drug: str) -> bool:
    """Match on drug_id (e.g. 'zoloft.139') or Drug column."""
    d = _drug(drug)
    # drug_id: "zoloft.139", "lexapro.22" …
    did = _n(row.get("drug_id", ""))
    if did.startswith(d + ".") or did == d:
        return True
    # "drug" / "Drug" column in Mapped sheets
    if _n(row.get("drug", row.get("Drug", ""))) == d:
        return True
    return False


def _match_cui(row: dict, cui: str) -> bool:
    """Match UMLS CUI in UMLS1/UMLS2 columns (compound: 'C0917801 / …')."""
    c = _n(cui)
    for key in ("UMLS1", "UMLS2"):
        val = _n(row.get(key, ""))
        if c in val:
            return True
    return False


def _match_text(row: dict, term: str) -> bool:
    """Substring match on all cell values."""
    t = _n(term)
    for v in row.values():
        if t in _n(v):
            return True
    return False


# ====================================================================
#  Core search
# ====================================================================

def search(entity: str, path: str = DATA_PATH,
           sheet: str | None = None,
           label: str | None = None) -> dict:
    """Search PsyTAR for *entity*.

    Parameters
    ----------
    entity : str
        Drug name (Zoloft / sertraline …), symptom (nausea …), or UMLS
        CUI (C0917801 …).
    sheet : str, optional
        Restrict to one sheet (fuzzy-matched, e.g. ``"ADR"`` →
        ``ADR_Identified``).
    label : str, optional
        In Sentence_Labeling, keep only rows where this label column
        (ADR / WD / EF / INF / SSI / DI) is positive (1).

    Returns
    -------
    dict : ``{sheet_name: [row_dict, …]}``
    """
    _load(path)
    etype = _entity_type(entity)
    matcher = {
        "drug": lambda row: _match_drug(row, entity),
        "cui":  lambda row: _match_cui(row, entity),
        "text": lambda row: _match_text(row, entity),
    }[etype]

    targets = list(_cache.keys())
    if sheet:
        r = _resolve(sheet)
        if r:
            targets = [r]

    # CUI queries only hit Mapped sheets (UMLS1/UMLS2 live there)
    if etype == "cui" and not sheet:
        targets = [s for s in targets if "mapped" in s.lower()]

    out: dict[str, list] = {}
    for sn in targets:
        hits = []
        for row in _cache[sn]:
            if not matcher(row):
                continue
            # label filter for Sentence_Labeling
            if label and "labeling" in sn.lower():
                flag = str(row.get(label.upper(),
                           row.get(label, ""))).strip()
                if flag in ("", "0", "nan", "None", "no"):
                    continue
            hits.append(row)
        if hits:
            out[sn] = hits
    return out


def search_batch(entities: list, path: str = DATA_PATH,
                 sheet: str | None = None,
                 label: str | None = None) -> dict:
    """Run :func:`search` for each entity.  Returns ``{entity: results}``."""
    _load(path)
    return {e: search(e, path, sheet, label) for e in entities}


# ====================================================================
#  Output
# ====================================================================

def summarize(results: dict, entity: str, max_rows: int = 10) -> str:
    """Compact LLM-readable summary."""
    if not results:
        return f"No PsyTAR results for '{entity}'."
    lines = [f"=== PsyTAR results for '{entity}' ==="]
    for sn, rows in results.items():
        lines.append(f"\n[{sn}]  ({len(rows)} hit{'s' if len(rows)!=1 else ''})")
        for r in rows[:max_rows]:
            parts = [f"{k}={str(v).strip()}" for k, v in r.items()
                     if str(v).strip() not in ("", "0", "nan", "None")]
            lines.append("  " + " | ".join(parts))
        if len(rows) > max_rows:
            lines.append(f"  ... ({len(rows)-max_rows} more)")
    return "\n".join(lines)


def to_json(results: dict) -> list:
    """Flatten to ``[{…, _sheet: …}]``."""
    return [dict(r, _sheet=sn) for sn, rows in results.items() for r in rows]


def describe(path: str = DATA_PATH) -> str:
    """Dataset overview."""
    _load(path)
    lines = ["=== PsyTAR Dataset ===", f"File: {path}"]
    for name, rows in _cache.items():
        cols = list(rows[0].keys()) if rows else []
        lines.append(f"  [{name}]  rows={len(rows)}  cols={cols}")
    lines.append("Drugs: Zoloft, Lexapro, Cymbalta, Effexor XR")
    lines.append("Labels: ADR · WD · SSI · DI · EF · INF · Findings · Others")
    return "\n".join(lines)


# ====================================================================
#  Demo
# ====================================================================

if __name__ == "__main__":
    import sys, json as _json
    if len(sys.argv) > 1:

        _cli_entities = sys.argv[1:]
        for _e in _cli_entities:
            _result = search(_e)
            print(summarize(_result, _e))
        sys.exit(0)

    # --- original demo below ---
    print(describe())

    # drug name (brand) → Mapped sheet
    res = search("Zoloft", sheet="ADR_Mapped")
    print("\n" + summarize(res, "Zoloft (ADR_Mapped)"))

    # generic name → same result
    res = search("sertraline", sheet="ADR_Mapped")
    print("\n" + summarize(res, "sertraline (ADR_Mapped)"))

    # symptom text → one Identified sheet
    res = search("nausea", sheet="ADR_Identified")
    print("\n" + summarize(res, "nausea (ADR_Identified)"))

    # symptom text → all sheets
    res = search("insomnia")
    print("\n" + summarize(res, "insomnia (all sheets)"))

    # UMLS CUI → auto-scoped to Mapped sheets
    res = search("C0917801")
    print("\n" + summarize(res, "C0917801 (auto→Mapped)"))

    # sentence-label filter: Effexor withdrawal sentences
    res = search("Effexor", sheet="Sentence_Labeling", label="WD")
    print("\n" + summarize(res, "Effexor (WD sentences)"))

    # batch
    batch = search_batch(["Lexapro", "insomnia", "C0917801"])
    for ent, hits in batch.items():
        print("\n" + summarize(hits, ent))

    # JSON output
    flat = to_json(search("Cymbalta", sheet="Sentence_Labeling"))
    print(f"\nJSON (Cymbalta sentences): {len(flat)}")
    if flat:
        print(json.dumps(flat[0], indent=2, default=str))