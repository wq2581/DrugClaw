"""
48_LTKB – FDA Liver Toxicity Knowledge Base (DILIrank + DILIst)
Category: Drug-centric | Type: Dataset | Subcategory: Drug Toxicity
Link: https://www.fda.gov/science-research/bioinformatics-tools/liver-toxicity-knowledge-base-ltkb
Paper: https://pubmed.ncbi.nlm.nih.gov/21624500/

Two datasets bundled:
  • DILIrank 2.0 – 1,036 drugs ranked by DILI potential (Most/Less/No/Ambiguous-concern)
  • DILIst       – 1,279 drugs classified by DILI severity (1 = positive, 0 = negative)

Access: Download Excel files from the FDA LTKB website.
"""

import os, json, re
from typing import Union

# ── Data paths (HPC absolute) ──────────────────────────────────────────
DATA_DIR = "/blue/qsong1/wang.qing/AgentLLM/DrugClaw/resources_metadata/drug_toxicity/DILI"
DILIRANK_FILE = os.path.join(DATA_DIR, "Drug Induced Liver Injury Rank (DILIrank 2.0) Dataset  FDA.xlsx")
DILIST_FILE   = os.path.join(DATA_DIR, "DILIst Supplementary Table.xlsx")


# ── Loaders ─────────────────────────────────────────────────────────────

def _read_excel(path: str) -> list[dict]:
    """Read an xlsx file into list[dict]. Uses openpyxl (stdlib-friendly)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("pip install openpyxl")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    # First non-None row as header
    header_idx = 0
    for i, row in enumerate(rows):
        if row and any(cell is not None for cell in row):
            header_idx = i
            break
    headers = [str(h).strip() if h else f"col_{j}" for j, h in enumerate(rows[header_idx])]
    records = []
    for row in rows[header_idx + 1:]:
        if not row or all(c is None for c in row):
            continue
        records.append({headers[j]: row[j] for j in range(min(len(headers), len(row)))})
    return records


def load_dilirank(path: str = DILIRANK_FILE) -> list[dict]:
    """Load DILIrank 2.0 dataset → list[dict].
    Keys: LTKBID, CompoundName, SeverityClass, LabelSection, vDILI-Concern, Comment
    """
    return _read_excel(path)


def load_dilist(path: str = DILIST_FILE) -> list[dict]:
    """Load DILIst dataset → list[dict].
    Keys: DILIST_ID, CompoundName, DILIst Classification, Routs of Administration
    """
    return _read_excel(path)


def load_all(dilirank_path: str = DILIRANK_FILE,
             dilist_path: str = DILIST_FILE) -> dict:
    """Load both datasets into {"dilirank": [...], "dilist": [...]}."""
    return {
        "dilirank": load_dilirank(dilirank_path),
        "dilist":   load_dilist(dilist_path),
    }


# ── Entity detection ────────────────────────────────────────────────────

def _detect_type(entity: str) -> str:
    """Classify entity string.
    LT\\d+     → ltkb_id   (DILIrank LTKB ID)
    digits     → dilist_id (DILIst numeric ID)
    otherwise  → drug_name (substring match)
    """
    e = entity.strip()
    if re.match(r"^LT\d+$", e, re.IGNORECASE):
        return "ltkb_id"
    if re.match(r"^\d+$", e):
        return "dilist_id"
    return "drug_name"


# ── Search ──────────────────────────────────────────────────────────────

def _search_dilirank(records: list[dict], entity: str, etype: str) -> list[dict]:
    hits = []
    low = entity.strip().lower()
    for r in records:
        if etype == "ltkb_id":
            if str(r.get("LTKBID", "")).strip().upper() == entity.strip().upper():
                hits.append(r)
        else:  # drug_name
            name = str(r.get("CompoundName", "")).lower()
            if low in name:
                hits.append(r)
    return hits


def _search_dilist(records: list[dict], entity: str, etype: str) -> list[dict]:
    hits = []
    low = entity.strip().lower()
    for r in records:
        if etype == "dilist_id":
            if str(r.get("DILIST_ID", "")).strip() == entity.strip():
                hits.append(r)
        else:  # drug_name
            name = str(r.get("CompoundName", "")).lower()
            if low in name:
                hits.append(r)
    return hits


def search(data: dict, entity: str) -> dict:
    """Search both DILIrank and DILIst for a single entity.

    Args:
        data: output of load_all()
        entity: drug name, LTKB ID (e.g. "LT00040"), or DILIst ID (e.g. "2")

    Returns:
        {"entity": str, "type": str,
         "dilirank": [matching records], "dilist": [matching records]}
    """
    etype = _detect_type(entity)
    return {
        "entity":   entity,
        "type":     etype,
        "dilirank": _search_dilirank(data["dilirank"], entity, etype) if etype != "dilist_id" else [],
        "dilist":   _search_dilist(data["dilist"], entity, etype)    if etype != "ltkb_id"   else [],
    }


def search_batch(data: dict, entities: list[str]) -> dict:
    """Search multiple entities. Returns {entity: search_result}."""
    return {e: search(data, e) for e in entities}


# ── Summarise (LLM-friendly) ───────────────────────────────────────────

_CONCERN_MAP = {
    "vMOST-DILI-concern":      "MOST-DILI-concern",
    "vLess-DILI-concern":      "Less-DILI-concern",
    "vNo-DILI-concern":        "No-DILI-concern",
    "vAmbiguous-DILI-concern": "Ambiguous-DILI-concern",
}

_DILIST_MAP = {1: "DILI-positive", 0: "DILI-negative", "1": "DILI-positive", "0": "DILI-negative"}


def summarize(result: dict, entity: str | None = None) -> str:
    """Compact one-line-per-hit summary for LLM context windows."""
    lines = []
    label = entity or result.get("entity", "?")

    for r in result.get("dilirank", []):
        concern = _CONCERN_MAP.get(str(r.get("vDILI-Concern", "")), str(r.get("vDILI-Concern", "")))
        sev = r.get("SeverityClass", "?")
        sec = r.get("LabelSection", "?")
        lines.append(f"DILIrank | {r.get('CompoundName','?')} | {concern} | severity={sev} | label_section={sec}")

    for r in result.get("dilist", []):
        cls_val = r.get("DILIst Classification", "?")
        cls_label = _DILIST_MAP.get(cls_val, str(cls_val))
        route = r.get("Routs of Administration", "?")
        lines.append(f"DILIst | {r.get('CompoundName','?')} | {cls_label} | route={route}")

    if not lines:
        return f"{label}: no LTKB records found."
    return f"{label}:\n" + "\n".join(f"  {l}" for l in lines)


def to_json(result: dict) -> str:
    """JSON serialisation of search result."""
    return json.dumps(result, default=str, ensure_ascii=False)


# ── Main (runnable examples) ───────────────────────────────────────────

if __name__ == "__main__":
    data = load_all()
    print(f"DILIrank: {len(data['dilirank'])} drugs | DILIst: {len(data['dilist'])} drugs\n")

    # 1) Single drug name search
    res = search(data, "acetaminophen")
    print(summarize(res))
    print()

    # 2) LTKB ID search (DILIrank only)
    res = search(data, "LT00040")
    print(summarize(res))
    print()

    # 3) DILIst numeric ID
    res = search(data, "2")
    print(summarize(res))
    print()

    # 4) Batch search
    batch = search_batch(data, ["aspirin", "methotrexate", "chlorpheniramine"])
    for entity, r in batch.items():
        print(summarize(r, entity))
    print()

    # 5) JSON output
    res = search(data, "abacavir")
    print(to_json(res))